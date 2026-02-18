"""
ETF 수익률 검증용 엑셀 내보내기
각 ETF별로 기준 날짜, 시작가, 종가, 계산 과정을 상세히 기록
"""

import yfinance as yf
import json
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

# JSON 데이터 로드
with open('../frontend/public/data/etf_database.json', 'r', encoding='utf-8') as f:
    database = json.load(f)

# 모든 ETF 티커 추출
etfs = []
for cat_key, cat in database['categories'].items():
    for etf in cat['etfs']:
        yf_ticker = f"{etf['ticker']}.KS" if etf['currency'] == 'KRW' else etf['ticker']
        etfs.append({
            'ticker': etf['ticker'],
            'yf_ticker': yf_ticker,
            'name': etf['name'],
            'currency': etf['currency'],
            'category': cat_key,
            'json_returns': etf['returns'],
            'json_cagr': etf.get('cagr', {}),
        })

wb = Workbook()

# ── Sheet 1: 수익률 검증 요약 ──
ws_summary = wb.active
ws_summary.title = "수익률 검증"

# 스타일 정의
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
green_font = Font(color="00B050", bold=True)
red_font = Font(color="FF0000", bold=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

headers = [
    'Ticker', 'Name', 'Category', 'Currency',
    '현재 종가', '현재일',
    '1M 기준일', '1M 기준일 종가', '1M 수익률(%)',
    '3M 기준일', '3M 기준일 종가', '3M 수익률(%)',
    '6M 기준일', '6M 기준일 종가', '6M 수익률(%)',
    '1Y 기준일', '1Y 기준일 종가', '1Y 수익률(%)',
    '3Y 기준일', '3Y 기준일 종가', '3Y 수익률(%)', '3Y CAGR(%)',
    '5Y 기준일', '5Y 기준일 종가', '5Y 수익률(%)', '5Y CAGR(%)',
]

for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

period_defs = [
    ('1M', relativedelta(months=1), 1/12),
    ('3M', relativedelta(months=3), 3/12),
    ('6M', relativedelta(months=6), 6/12),
    ('1Y', relativedelta(years=1), 1),
    ('3Y', relativedelta(years=3), 3),
    ('5Y', relativedelta(years=5), 5),
]

print(f"총 {len(etfs)}개 ETF 검증 데이터 생성 중...\n")

for idx, etf_info in enumerate(etfs):
    row = idx + 2
    ticker = etf_info['yf_ticker']
    print(f"  [{idx+1}/{len(etfs)}] {etf_info['ticker']}...", end=' ')

    try:
        etf = yf.Ticker(ticker)
        history = etf.history(period="10y")

        if len(history) == 0:
            print("❌ 데이터 없음")
            continue

        current_price = history['Close'].iloc[-1]
        last_date = history.index[-1]

        ws_summary.cell(row=row, column=1, value=etf_info['ticker']).border = thin_border
        ws_summary.cell(row=row, column=2, value=etf_info['name']).border = thin_border
        ws_summary.cell(row=row, column=3, value=etf_info['category']).border = thin_border
        ws_summary.cell(row=row, column=4, value=etf_info['currency']).border = thin_border
        c = ws_summary.cell(row=row, column=5, value=round(current_price, 2))
        c.number_format = '#,##0.00'
        c.border = thin_border
        ws_summary.cell(row=row, column=6, value=last_date.strftime('%Y-%m-%d')).border = thin_border

        col_offset = 7
        for p_name, delta, years in period_defs:
            target_date = last_date - delta
            mask = history.index <= target_date
            if mask.any():
                start_row = history.loc[mask].iloc[-1]
                start_date = history.index[mask][-1]
                start_price = start_row['Close']
                cumulative = ((current_price / start_price) - 1) * 100

                ws_summary.cell(row=row, column=col_offset, value=start_date.strftime('%Y-%m-%d')).border = thin_border
                c = ws_summary.cell(row=row, column=col_offset + 1, value=round(start_price, 2))
                c.number_format = '#,##0.00'
                c.border = thin_border
                c = ws_summary.cell(row=row, column=col_offset + 2, value=round(cumulative, 2))
                c.number_format = '0.00'
                c.font = green_font if cumulative >= 0 else red_font
                c.border = thin_border

                if years >= 3:
                    cagr_val = ((current_price / start_price) ** (1 / years) - 1) * 100
                    c = ws_summary.cell(row=row, column=col_offset + 3, value=round(cagr_val, 2))
                    c.number_format = '0.00'
                    c.font = green_font if cagr_val >= 0 else red_font
                    c.border = thin_border
                    col_offset += 4
                else:
                    col_offset += 3
            else:
                for i in range(3 if years < 3 else 4):
                    ws_summary.cell(row=row, column=col_offset + i, value='-').border = thin_border
                col_offset += 3 if years < 3 else 4

        print("✅")

    except Exception as e:
        print(f"❌ {str(e)[:40]}")

# ── Sheet 2: 주요 ETF 일별 가격 (최근 1년) ──
ws_daily = wb.create_sheet("일별 가격 (1년)")
sample_tickers = ['SPY', 'QQQ', 'SCHD', 'GLD', 'TLT']

daily_headers = ['Date'] + [f'{t} Close' for t in sample_tickers]
for col, h in enumerate(daily_headers, 1):
    cell = ws_daily.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

print(f"\n주요 ETF 일별 가격 수집 중 ({', '.join(sample_tickers)})...")
histories = {}
for t in sample_tickers:
    histories[t] = yf.Ticker(t).history(period="1y")

# 날짜 합치기
all_dates = set()
for h in histories.values():
    all_dates.update(h.index)
all_dates = sorted(all_dates)

for row_idx, date in enumerate(all_dates, 2):
    ws_daily.cell(row=row_idx, column=1, value=date.strftime('%Y-%m-%d')).border = thin_border
    for col_idx, t in enumerate(sample_tickers, 2):
        if date in histories[t].index:
            val = round(histories[t].loc[date, 'Close'], 2)
            c = ws_daily.cell(row=row_idx, column=col_idx, value=val)
            c.number_format = '#,##0.00'
        else:
            c = ws_daily.cell(row=row_idx, column=col_idx, value='')
        c.border = thin_border

# ── Sheet 3: 계산 방법 설명 ──
ws_method = wb.create_sheet("계산 방법")
explanations = [
    ['항목', '설명'],
    ['데이터 소스', 'Yahoo Finance (yfinance 라이브러리)'],
    ['가격 기준', 'Adjusted Close (배당 재투자 + 분할 조정 반영)'],
    ['수익률 유형', 'Total Return (배당 포함)'],
    ['기간 계산 방식', '달력 날짜 기준 (예: 1Y = 정확히 1년 전)'],
    ['기준일 선택', '목표 날짜 이전 마지막 거래일 (Yahoo Finance 방식)'],
    ['수익률 공식', '((현재가 / 시작가) - 1) × 100'],
    ['CAGR 공식', '((현재가 / 시작가) ^ (1/연수) - 1) × 100'],
    ['변동성', '일별 수익률의 표준편차 × √252 (연환산)'],
    ['최대 낙폭 (MDD)', '누적 수익률 대비 고점에서 최대 하락폭'],
    ['', ''],
    ['기간별 기준', ''],
    ['1M', '1개월 전 달력 날짜'],
    ['3M', '3개월 전 달력 날짜'],
    ['6M', '6개월 전 달력 날짜'],
    ['1Y', '1년 전 달력 날짜'],
    ['3Y', '3년 전 달력 날짜 (CAGR 포함)'],
    ['5Y', '5년 전 달력 날짜 (CAGR 포함)'],
    ['', ''],
    ['참고', 'Yahoo Finance, Barchart 등과 소폭 차이는 데이터 수집 시점 차이에 의한 것'],
    ['수집 시점', database['metadata']['lastUpdate']],
]

for row_idx, (a, b) in enumerate(explanations, 1):
    c1 = ws_method.cell(row=row_idx, column=1, value=a)
    c2 = ws_method.cell(row=row_idx, column=2, value=b)
    if row_idx == 1:
        c1.font = header_font
        c1.fill = header_fill
        c2.font = header_font
        c2.fill = header_fill
    c1.border = thin_border
    c2.border = thin_border

# 열 너비 조정
for ws in [ws_summary, ws_daily, ws_method]:
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)

# 저장
output_path = '../ETF_검증데이터_v2.xlsx'
wb.save(output_path)
print(f"\n✅ 엑셀 저장 완료: {output_path}")
print("  - Sheet 1: 수익률 검증 (전체 54개 ETF 시작가/종가/수익률)")
print("  - Sheet 2: 일별 가격 (SPY/QQQ/SCHD/GLD/TLT 최근 1년)")
print("  - Sheet 3: 계산 방법 설명")
